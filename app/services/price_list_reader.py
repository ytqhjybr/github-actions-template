import openpyxl
from typing import List, Dict, Any

def read_price_list(file_path: str) -> List[Dict[str, Any]]:
    """
    Читает Excel-файл и возвращает список словарей.
    Пытается определить колонки: если есть "Наименование" или "Товар", то использует как name.
    Если есть "Цена", "Price" или "Стоимость", то как price.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Получаем заголовки
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        headers.append(str(cell.value).strip() if cell.value else f"col_{col}")

    # Ищем колонки
    name_col = None
    price_col = None
    for i, h in enumerate(headers):
        if h.lower() in ["наименование", "товар", "название", "name"]:
            name_col = i
        if h.lower() in ["цена", "price", "стоимость", "cost"]:
            price_col = i

    # Если не нашли, берём первые две колонки
    if name_col is None:
        name_col = 0
    if price_col is None and len(headers) > 1:
        price_col = 1

    data = []
    for row in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers):
            cell_value = ws.cell(row=row, column=col_idx+1).value
            if cell_value is not None:
                row_data[header] = cell_value
        # Добавляем "умолчательные" ключи для единого доступа
        if name_col is not None:
            row_data["name"] = ws.cell(row=row, column=name_col+1).value
        if price_col is not None:
            row_data["price"] = ws.cell(row=row, column=price_col+1).value
        if row_data.get("name") or row_data.get("price"):
            data.append(row_data)

    return data